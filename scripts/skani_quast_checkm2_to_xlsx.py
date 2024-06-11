#!/usr/bin/python3
import os
import sys
import csv
from openpyxl import Workbook

#import the location of the argument line, this location comes from the bashscript scripts/complete_illuminapipeline.sh
#If you want to use this script without the bash script you can use the following command line
#location ="/home/genomics/mhannaert/data/mini_testdata/gz_files/output_test4/"

location =  sys.argv[1]

#going to the location for the output file
os.chdir(location)
print("\nThe output file can be found in: ", os.getcwd())

#opening a xlsx document 
wb = Workbook() 
ws = wb.active 

#making a first sheet for skANI
ws.title = "skANI_output"
with open('06_skani/skani_results_file.txt') as csv_file:
    csv_reader = csv.reader(csv_file, delimiter='\t')
    line_count = 0
    for row in csv_reader:
        ws.append(row)
        line_count += 1
    print("Processed {} lines.".format(line_count))
csv_file.close()

#making a second sheet for Quast
ws2 = wb.create_sheet(title="Quast_output")
with open('07_quast/quast_summary_table.txt') as csv_file:
    csv_reader = csv.reader(csv_file, delimiter='\t')
    line_count = 0
    for row in csv_reader:
        ws2.append(row)
        line_count += 1
    print("Processed {} lines.".format(line_count))
csv_file.close()

#making a second sheet for Quast
ws3 = wb.create_sheet(title="CheckM2_output")
with open('09_checkM2/checkM2_summary_table.txt') as csv_file:
    csv_reader = csv.reader(csv_file, delimiter='\t')
    line_count = 0
    for row in csv_reader:
        ws3.append(row)
        line_count += 1
    print("Processed {} lines.".format(line_count))
csv_file.close()

#closing the workbook and saving it with the following name
wb.save("skANI_Quast_checkM2_output.xlsx")
print("\nskANI_Quast_checkM2_output.xlsx is made in: ", os.getcwd())